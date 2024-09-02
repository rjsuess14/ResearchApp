import os
from dotenv import load_dotenv
import fmpsdk
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import requests
# Actual API key is stored in a .env file.  Not good to store API key directly in script.
load_dotenv()
apikey = os.environ.get("FMP_KEY")


#Fetch financial statement data
def fmp_fs_data(ticker, period):
    is_url = f"https://financialmodelingprep.com/api/v3/income-statement/{ticker}?period={period}&limit=400&apikey={apikey}"
    is_response = requests.get(is_url)
    bs_url = f"https://financialmodelingprep.com/api/v3/balance-sheet-statement/{ticker}?period={period}&limit=400&apikey={apikey}"
    bs_response = requests.get(bs_url)
    cfs_url = f"https://financialmodelingprep.com/api/v3/cash-flow-statement/{ticker}?period={period}&limit=400&apikey={apikey}"
    cfs_response = requests.get(cfs_url)

    if is_response.status_code == 200 and bs_response.status_code == 200 and cfs_response.status_code == 200:
        is_data = is_response.json()
        fmp_is_df = pd.DataFrame(is_data)
        bs_data = bs_response.json()
        fmp_bs_df = pd.DataFrame(bs_data)
        cfs_data = cfs_response.json()
        fmp_cfs_df = pd.DataFrame(cfs_data)
    else:
        print("Failed to retrieve data from the API")
    return fmp_is_df, fmp_bs_df, fmp_cfs_df

#Fetch company info

def fetch_company_info(ticker):
    """
    Fetch company information from the FMP API and return as a DataFrame.

    Parameters:
    ticker (str): The stock ticker symbol of the company.

    Returns:
    pd.DataFrame: A DataFrame containing company information.
    """
    url = f"https://financialmodelingprep.com/api/v3/profile/{ticker}?apikey={apikey}"
    response = requests.get(url)

    if response.status_code == 200:
        profile_data = response.json()
        if profile_data:
            # Convert the dictionary to a DataFrame
            df = pd.DataFrame([profile_data[0]])
            return df
        else:
            print(f"No data found for ticker: {ticker}")
            return pd.DataFrame()  # Return an empty DataFrame
    else:
        print(f"Failed to retrieve data from the API. Status code: {response.status_code}")
        return pd.DataFrame()  # Return an empty DataFrame




def save_to_excel(is_df, bs_df, cfs_df, filename):
    wb = Workbook()
    
    # Create sheets
    ws1 = wb.active
    ws1.title = "Income Statement"
    ws2 = wb.create_sheet(title="Balance Sheet")
    ws3 = wb.create_sheet(title="Cash Flow Statement")
    
    # Write DataFrames to sheets
    for row in dataframe_to_rows(is_df, index=False, header=True):
        ws1.append(row)
    
    for row in dataframe_to_rows(bs_df, index=False, header=True):
        ws2.append(row)
    
    for row in dataframe_to_rows(cfs_df, index=False, header=True):
        ws3.append(row)
    
    # Save the workbook
    wb.save(filename)

# Usage
is_df, bs_df, cfs_df = fmp_fs_data('NSRGY', 'annual')
save_to_excel(is_df, bs_df, cfs_df, f'fmp_output.xlsx')

