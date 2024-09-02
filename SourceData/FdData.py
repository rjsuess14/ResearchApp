import os
from dotenv import load_dotenv
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import requests

# Actual API key is stored in a .env file.  Not good to store API key directly in script.
load_dotenv()
apikey = os.environ.get("FDAI_KEY")

headers = {
    "X-API-KEY": apikey
}

#Fetch financial statement data
def fd_fs_data(ticker, period, limit):
    fs_url = f'https://api.financialdatasets.ai/financials/'
    querystring = {"ticker":ticker,"period":period,"limit":limit}
    fs_response = requests.get(fs_url, headers=headers, params=querystring)

    if fs_response.status_code == 200:
        fs_data = fs_response.json().get('financials')
        is_data = fs_data.get('income_statements')
        bs_data = fs_data.get('balance_sheets')
        cfs_data = fs_data.get('cash_flow_statements')

        fd_is_df = pd.DataFrame(is_data)
        fd_bs_df = pd.DataFrame(bs_data)
        fd_cfs_df = pd.DataFrame(cfs_data)

    else:
        print("Failed to retrieve data from the API")
    return fd_is_df, fd_bs_df, fd_cfs_df

#Fetch financial statement data
def price_data(ticker, interval, multiplier, start_date, end_date):
    price_url = f'https://api.financialdatasets.ai/prices'
    querystring = {"ticker":ticker,"interval":interval,"interval_multiplier":multiplier,"start_date":start_date,"end_date":end_date}
    price_response = requests.get(price_url, headers=headers, params=querystring)

    if price_response.status_code == 200:
        price_data = price_response.json().get('prices')

        price_df = pd.DataFrame(price_data)

    else:
        print("Failed to retrieve data from the API")
    return price_df

def save_to_excel(is_df, bs_df, cfs_df, price_df, filename):
    wb = Workbook()
    
    # Create sheets
    ws1 = wb.active
    ws1.title = "Income Statement"
    ws2 = wb.create_sheet(title="Balance Sheet")
    ws3 = wb.create_sheet(title="Cash Flow Statement")
    ws4 = wb.create_sheet(title="Price History")
    
    # Write DataFrames to sheets
    for row in dataframe_to_rows(is_df, index=False, header=True):
        ws1.append(row)
    
    for row in dataframe_to_rows(bs_df, index=False, header=True):
        ws2.append(row)
    
    for row in dataframe_to_rows(cfs_df, index=False, header=True):
        ws3.append(row)
    
    for row in dataframe_to_rows(price_df, index=False, header=True):
        ws4.append(row)
    # Save the workbook
    wb.save(filename)

# Usage
ticker_input = "BIRK"
is_df, bs_df, cfs_df = fd_fs_data(ticker_input, 'quarterly', 8)
price_df = price_data(ticker_input, 'quarter', 1, '2019-01-01', '2024-06-30')
save_to_excel(is_df, bs_df, cfs_df, price_df, f'{ticker_input}_output.xlsx')

